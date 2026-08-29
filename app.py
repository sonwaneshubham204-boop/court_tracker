import os
import uuid
import re

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    send_file,
    send_from_directory,
    session,
    flash,
    url_for
)

from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, date, timedelta
from openpyxl import Workbook, load_workbook
from io import BytesIO
from markupsafe import Markup, escape


app = Flask(__name__)

app.config["SECRET_KEY"] = os.environ.get(
    "SECRET_KEY",
    "change-this-to-a-long-random-secret-key"
)


# =========================================================
# DATABASE CONFIGURATION
# =========================================================

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "sqlite:///court_tracker.db"
)

# Some hosting platforms provide postgres://; SQLAlchemy expects postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace(
        "postgres://",
        "postgresql://",
        1
    )

app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False


# =========================================================
# FILE UPLOAD CONFIGURATION
# =========================================================

BASE_DIR = os.path.abspath(
    os.path.dirname(__file__)
)

UPLOAD_FOLDER = os.path.join(
    BASE_DIR,
    "uploads"
)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER


os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)


ALLOWED_EXTENSIONS = {

    "pdf",
    "doc",
    "docx",
    "txt",
    "xlsx",

    "png",
    "jpg",
    "jpeg",
    "gif"
}


db = SQLAlchemy(app)

# =========================================================
# PARTY NAME DISPLAY FILTER
# Only "Vs" / "vs" is bold and chocolate coloured.
# =========================================================

@app.template_filter("format_vs")
def format_vs(value):
    text = str(value or "")
    parts = re.split(r"((?<!\w)vs\.?(?!\w))", text, flags=re.IGNORECASE)
    return Markup("".join(
        '<span class="vs-highlight">' + str(escape(part)) + '</span>'
        if re.fullmatch(r"vs\.?", part, flags=re.IGNORECASE)
        else str(escape(part))
        for part in parts
    ))


def natural_case_key(case):
    """Sort case numbers by their text and numeric parts, not as plain strings."""
    return tuple(
        (0, int(part)) if part.isdigit() else (1, part.casefold())
        for part in re.split(r"(\d+)", case.case_no or "")
    )


def sort_cases_naturally(cases):
    return sorted(cases, key=natural_case_key)


# =========================================================
# USER MODEL
# =========================================================

class User(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    username = db.Column(
        db.String(100),
        unique=True,
        nullable=False
    )

    password_hash = db.Column(
        db.String(255),
        nullable=False
    )


# =========================================================
# HELPER FUNCTION
# =========================================================

def allowed_file(filename):

    return (
        "." in filename
        and filename.rsplit(
            ".",
            1
        )[1].lower()
        in ALLOWED_EXTENSIONS
    )

# =========================================================
# LOGIN REQUIRED DECORATOR
# =========================================================

def login_required(view):

    @wraps(view)
    def wrapped_view(*args, **kwargs):

        if "user_id" not in session:

            flash(
                "Please log in to continue.",
                "warning"
            )

            return redirect(
                url_for("login")
            )

        return view(*args, **kwargs)

    return wrapped_view

# =========================================================
# CASE DATABASE MODEL
# =========================================================

class Case(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    case_no = db.Column(
        db.String(100),
        nullable=False
    )

    crn_no = db.Column(
        db.String(100),
        nullable=True
    )

    advocate_name = db.Column(
        db.String(200),
        nullable=True
    )

    case_disposed = db.Column(
        db.String(20),
        nullable=True,
        default="No"
    )

    court_no = db.Column(
        db.Integer,
        nullable=False
    )

    parties = db.Column(
        db.String(200),
        nullable=False
    )

    case_stage = db.Column(
        db.String(100),
        nullable=False
    )

    case_stage_other = db.Column(
        db.String(200),
        nullable=True
    )

    next_hearing_date = db.Column(
        db.Date,
        nullable=True
    )

    decision_date = db.Column(
        db.Date,
        nullable=True
    )

    highlighted = db.Column(
        db.Boolean,
        nullable=False,
        default=False
    )


    hearings = db.relationship(
        "Hearing",
        backref="case",
        lazy=True,
        cascade="all, delete-orphan"
    )


    evidences = db.relationship(
        "Evidence",
        backref="case",
        lazy=True,
        cascade="all, delete-orphan"
    )


# =========================================================
# INDEPENDENT ENQUIRY MODULE
# =========================================================

class Enquiry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    organisation_name = db.Column(db.String(250), nullable=False)
    start_date = db.Column(db.Date, nullable=False)
    end_date = db.Column(db.Date, nullable=True)
    next_enquiry_date = db.Column(db.Date, nullable=True)
    next_enquiry_time = db.Column(db.Time, nullable=True)
    status = db.Column(db.String(50), nullable=False, default="Pending for Report")
    highlighted = db.Column(db.Boolean, nullable=False, default=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    sittings = db.relationship("EnquirySitting", backref="enquiry", lazy=True, cascade="all, delete-orphan")

    @property
    def total_turns(self):
        return len(self.sittings)


class EnquirySitting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    enquiry_id = db.Column(db.Integer, db.ForeignKey("enquiry.id"), nullable=False)
    sitting_date = db.Column(db.Date, nullable=False)
    sitting_time = db.Column(db.Time, nullable=True)
    remark = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)


# =========================================================
# HEARING MODEL
# =========================================================

class Hearing(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    case_id = db.Column(
        db.Integer,
        db.ForeignKey("case.id"),
        nullable=False
    )

    hearing_date = db.Column(
        db.Date,
        nullable=False,
        default=date.today
    )

    outcome = db.Column(
        db.String(200),
        nullable=False
    )

    presentee = db.Column(
        db.Text,
        nullable=True
    )

    business = db.Column(
        db.Text,
        nullable=True
    )

    next_hearing_date = db.Column(
        db.Date,
        nullable=True
    )

    notes = db.Column(
        db.Text,
        nullable=True
    )


# =========================================================
# EVIDENCE MODEL
# =========================================================

class Evidence(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    case_id = db.Column(
        db.Integer,
        db.ForeignKey("case.id"),
        nullable=False
    )

    evidence_name = db.Column(
        db.String(200),
        nullable=False
    )

    evidence_type = db.Column(
        db.String(100),
        nullable=False
    )

    description = db.Column(
        db.Text,
        nullable=True
    )

    submitted_date = db.Column(
        db.Date,
        nullable=True
    )

    status = db.Column(
        db.String(100),
        nullable=False,
        default="Pending"
    )


    files = db.relationship(
        "EvidenceFile",
        backref="evidence",
        lazy=True,
        cascade="all, delete-orphan"
    )


# =========================================================
# MULTIPLE EVIDENCE FILE MODEL
# =========================================================

class EvidenceFile(db.Model):

    id = db.Column(
        db.Integer,
        primary_key=True
    )

    evidence_id = db.Column(
        db.Integer,
        db.ForeignKey("evidence.id"),
        nullable=False
    )

    file_name = db.Column(
        db.String(300),
        nullable=False
    )

    original_filename = db.Column(
        db.String(300),
        nullable=False
    )

# =========================================================
# LOGIN
# =========================================================

@app.route("/login", methods=["GET", "POST"])
def login():

    if "user_id" in session:
        return redirect(url_for("home"))

    if request.method == "POST":

        username = request.form.get(
            "username"
        ).strip()

        password = request.form.get(
            "password"
        )

        user = User.query.filter_by(
            username=username
        ).first()

        if user and check_password_hash(
            user.password_hash,
            password
        ):

            session["user_id"] = user.id

            flash(
                "Login successful!",
                "success"
            )

            return redirect(
                url_for("home")
            )

        flash(
            "Invalid username or password.",
            "danger"
        )

    return render_template(
        "login.html"
    )


# =========================================================
# LOGOUT
# =========================================================

@app.route("/logout")
def logout():

    session.clear()

    flash(
        "You have been logged out.",
        "success"
    )

    return redirect(
        url_for("login")
    )

# =========================================================
# SMART DASHBOARD
# =========================================================

@app.route("/")
@login_required
def home():

    today = date.today()

    tomorrow = today + timedelta(
        days=1
    )

    next_7_days = today + timedelta(
        days=7
    )

    selected_matter_date = today
    selected_matter_date_value = request.args.get("matter_date", "").strip()
    if selected_matter_date_value:
        try:
            selected_matter_date = datetime.strptime(
                selected_matter_date_value, "%Y-%m-%d"
            ).date()
        except ValueError:
            pass


    total_cases = Case.query.count()


    court_1_count = Case.query.filter_by(
        court_no=1
    ).count()

    court_2_count = Case.query.filter_by(
        court_no=2
    ).count()

    court_3_count = Case.query.filter_by(
        court_no=3
    ).count()


    past_cases = Case.query.filter(
        Case.next_hearing_date < today
    ).count()


    todays_cases = Case.query.filter(
        Case.next_hearing_date == today
    ).count()


    tomorrow_cases = Case.query.filter(
        Case.next_hearing_date == tomorrow
    ).count()


    upcoming_cases = Case.query.filter(
        Case.next_hearing_date > today
    ).count()


    next_7_days_count = Case.query.filter(
        Case.next_hearing_date >= today,
        Case.next_hearing_date <= next_7_days
    ).count()


    undated_cases = Case.query.filter(
        Case.next_hearing_date.is_(None),
        Case.decision_date.is_(None),
        Case.case_disposed != "Yes"
    ).count()


    disposed_cases = Case.query.filter_by(
        case_disposed="Yes"
    ).count()

    # =====================================================
    # ENQUIRY DASHBOARD COUNTS
    # =====================================================
    _ensure_enquiry_tables()

    todays_enquiries = Enquiry.query.filter(
        Enquiry.next_enquiry_date == today
    ).count()

    next_enquiries = Enquiry.query.filter(
        Enquiry.next_enquiry_date > today
    ).count()


    # =====================================================
    # CASE STAGE COUNTS
    # =====================================================

    notice_count = Case.query.filter_by(
        case_stage="Notice"
    ).count()

    admission_count = Case.query.filter_by(
        case_stage="Admission"
    ).count()

    hearing_count = Case.query.filter_by(
        case_stage="Hearing"
    ).count()

    evidence_count = Case.query.filter_by(
        case_stage="Evidence"
    ).count()

    cross_examination_count = Case.query.filter_by(
        case_stage="Cross Examination"
    ).count()

    arguments_count = Case.query.filter_by(
        case_stage="Arguments"
    ).count()

    final_arguments_count = Case.query.filter_by(
        case_stage="Final Arguments"
    ).count()

    final_hearing_count = Case.query.filter_by(
        case_stage="Final Hearing"
    ).count()

    judgment_count = Case.query.filter_by(
        case_stage="Judgment"
    ).count()

    order_count = Case.query.filter_by(
        case_stage="Order"
    ).count()

    court_stage_counts = {court_no: {} for court_no in (1, 2, 3)}
    for case in Case.query.all():
        stage = (
            case.case_stage_other.strip()
            if case.case_stage == "Other" and case.case_stage_other
            else (case.case_stage or "").strip()
        )
        if stage and case.court_no in court_stage_counts:
            counts = court_stage_counts[case.court_no]
            counts[stage] = counts.get(stage, 0) + 1

    court_stage_counts = {
        court_no: sorted(counts.items(), key=lambda item: item[0].casefold())
        for court_no, counts in court_stage_counts.items()
    }

    selected_date_court_counts = {
        court_no: Case.query.filter_by(
            court_no=court_no,
            next_hearing_date=selected_matter_date
        ).count()
        for court_no in (1, 2, 3)
    }
    selected_date_total = sum(selected_date_court_counts.values())

    # Case stages represented by matters scheduled on the selected date,
    # grouped by court. This is used only for the dashboard visual highlight.
    selected_date_stage_counts = {court_no: {} for court_no in (1, 2, 3)}
    selected_date_cases = Case.query.filter(
        Case.next_hearing_date == selected_matter_date
    ).all()
    for case in selected_date_cases:
        stage = (
            case.case_stage_other.strip()
            if case.case_stage == "Other" and case.case_stage_other
            else (case.case_stage or "").strip()
        )
        if stage and case.court_no in selected_date_stage_counts:
            counts = selected_date_stage_counts[case.court_no]
            counts[stage] = counts.get(stage, 0) + 1

    selected_date_stage_counts = {
        court_no: set(counts.keys())
        for court_no, counts in selected_date_stage_counts.items()
    }


    # =====================================================
    # REMINDER LISTS
    # =====================================================

    todays_hearings = sort_cases_naturally(Case.query.filter(
        Case.next_hearing_date == today
    ).all())


    tomorrow_hearings = sort_cases_naturally(Case.query.filter(
        Case.next_hearing_date == tomorrow
    ).all())


    upcoming_hearings = Case.query.filter(
        Case.next_hearing_date > tomorrow,
        Case.next_hearing_date <= next_7_days
    ).all()
    upcoming_hearings = sort_cases_naturally(upcoming_hearings)


    overdue_hearings = Case.query.filter(
        Case.next_hearing_date < today
    ).all()
    overdue_hearings = sort_cases_naturally(overdue_hearings)


    # =====================================================
    # RECENT HEARING UPDATES
    # =====================================================

    recent_hearings = Hearing.query.order_by(
        Hearing.id.desc()
    ).limit(5).all()


    # =====================================================
    # RECENT EVIDENCE
    # =====================================================

    recent_evidences = Evidence.query.order_by(
        Evidence.id.desc()
    ).limit(5).all()


    return render_template(

        "dashboard.html",

        total_cases=total_cases,

        court_1_count=court_1_count,
        court_2_count=court_2_count,
        court_3_count=court_3_count,

        past_cases=past_cases,
        todays_cases=todays_cases,
        tomorrow_cases=tomorrow_cases,
        upcoming_cases=upcoming_cases,
        next_7_days_count=next_7_days_count,
        undated_cases=undated_cases,
        disposed_cases=disposed_cases,

        todays_enquiries=todays_enquiries,
        next_enquiries=next_enquiries,

        notice_count=notice_count,
        admission_count=admission_count,
        hearing_count=hearing_count,
        evidence_count=evidence_count,
        cross_examination_count=cross_examination_count,
        arguments_count=arguments_count,
        final_arguments_count=final_arguments_count,
        final_hearing_count=final_hearing_count,
        judgment_count=judgment_count,
        order_count=order_count,
        court_stage_counts=court_stage_counts,
        selected_date_stage_counts=selected_date_stage_counts,
        selected_matter_date=selected_matter_date,
        selected_date_court_counts=selected_date_court_counts,
        selected_date_total=selected_date_total,

        todays_hearings=todays_hearings,
        tomorrow_hearings=tomorrow_hearings,
        upcoming_hearings=upcoming_hearings,
        overdue_hearings=overdue_hearings,

        recent_hearings=recent_hearings,
        recent_evidences=recent_evidences,

        today=today,
        tomorrow=tomorrow
    )


# =========================================================
# ADVANCED SEARCH + FILTER
# =========================================================

@app.route("/cases")
@login_required
def case_list():

    search = request.args.get(
        "search",
        ""
    ).strip()

    court_no = request.args.get(
        "court_no",
        ""
    ).strip()

    case_stage = request.args.get(
        "case_stage",
        ""
    ).strip()

    hearing_status = request.args.get(
        "hearing_status",
        ""
    ).strip()

    case_disposed = request.args.get(
        "case_disposed",
        ""
    ).strip()

    highlighted_only = request.args.get(
        "highlighted",
        ""
    ).strip()

    date_from = request.args.get(
        "date_from",
        ""
    ).strip()

    date_to = request.args.get(
        "date_to",
        ""
    ).strip()


    today = date.today()


    query = Case.query


    # SEARCH CASE NUMBER / PARTY NAME
    # Numeric search must return only the case whose CASE NUMBER contains
    # that complete numeric segment; it must not search hidden CNR/party/advocate fields.
    if search:
        if search.strip().isdigit():
            term = search.strip()
            query = query.filter(
                db.or_(
                    Case.case_no == term,
                    Case.case_no.ilike(f"%/{term}/%"),
                    Case.case_no.ilike(f"% {term} %"),
                    Case.case_no.ilike(f"% {term}/%"),
                    Case.case_no.ilike(f"%/{term} %")
                )
            )
        else:
            query = query.filter(
                db.or_(
                    Case.case_no.ilike(f"%{search}%"),
                    Case.parties.ilike(f"%{search}%"),
                    Case.crn_no.ilike(f"%{search}%"),
                    Case.advocate_name.ilike(f"%{search}%")
                )
            )

    if court_no:

        try:

            query = query.filter(
                Case.court_no == int(
                    court_no
                )
            )

        except ValueError:

            pass


    # CASE STAGE FILTER
    # Match both normal stages and the actual custom stage saved under Other.
    if case_stage:
        query = query.filter(
            db.or_(
                Case.case_stage == case_stage,
                db.and_(
                    Case.case_stage == "Other",
                    Case.case_stage_other == case_stage
                )
            )
        )


    # HEARING STATUS FILTER

    if hearing_status == "past":

        query = query.filter(
            Case.next_hearing_date < today
        )


    elif hearing_status == "today":

        query = query.filter(
            Case.next_hearing_date == today
        )


    elif hearing_status == "tomorrow":

        query = query.filter(
            Case.next_hearing_date == (
                today + timedelta(days=1)
            )
        )


    elif hearing_status == "upcoming":

        query = query.filter(
            Case.next_hearing_date > today
        )


    elif hearing_status == "undated":

        query = query.filter(
            Case.next_hearing_date.is_(None),
            Case.decision_date.is_(None),
            Case.case_disposed != "Yes"
        )


    # HIGHLIGHTED CASES FILTER
    if highlighted_only == "1":
        query = query.filter(Case.highlighted.is_(True))

    # CASE DISPOSED FILTER

    if case_disposed in ["Yes", "No"]:

        query = query.filter(
            Case.case_disposed == case_disposed
        )


    # FROM DATE FILTER

    if date_from:

        try:

            from_date = datetime.strptime(
                date_from,
                "%Y-%m-%d"
            ).date()

            query = query.filter(
                Case.next_hearing_date >= from_date
            )

        except ValueError:

            pass


    # TO DATE FILTER

    if date_to:

        try:

            to_date = datetime.strptime(
                date_to,
                "%Y-%m-%d"
            ).date()

            query = query.filter(
                Case.next_hearing_date <= to_date
            )

        except ValueError:

            pass


    cases = sort_cases_naturally(query.all())

    # Build the Case Stage dropdown from every stage actually stored in the database.
    # For "Other", show the custom value when one exists.
    all_stage_options = set()
    for stage_value, custom_stage_value in db.session.query(
        Case.case_stage, Case.case_stage_other
    ).all():
        if stage_value == "Other" and custom_stage_value and custom_stage_value.strip():
            all_stage_options.add(custom_stage_value.strip())
        elif stage_value and stage_value.strip():
            all_stage_options.add(stage_value.strip())

    all_stage_options = sorted(all_stage_options, key=str.casefold)


    return render_template(

        "cases.html",

        cases=cases,

        today=today,

        title="Court Cases",

        search=search,

        court_no=court_no,

        case_stage=case_stage,

        hearing_status=hearing_status,

        case_disposed=case_disposed,

        highlighted_only=highlighted_only,

        date_from=date_from,

        date_to=date_to,

        all_stage_options=all_stage_options
    )


# =========================================================
# TOGGLE CASE HIGHLIGHT
# =========================================================

@app.route("/case/<int:id>/toggle-highlight", methods=["POST"])
@login_required
def toggle_case_highlight(id):
    case = Case.query.get_or_404(id)
    case.highlighted = not bool(case.highlighted)
    db.session.commit()
    return redirect(request.referrer or url_for("case_list"))


# =========================================================
# COURT-WISE CASES
# =========================================================

@app.route("/court/<int:court_no>")
@login_required
def court_cases(court_no):

    if court_no not in [1, 2, 3]:

        return "Invalid Court Number"


    cases = sort_cases_naturally(Case.query.filter_by(
        court_no=court_no
    ).all())


    return render_template(

        "cases.html",

        cases=cases,

        today=date.today(),

        title=f"Court {court_no} Cases",

        search="",

        court_no=str(court_no),

        case_stage="",

        hearing_status="",

        highlighted_only="",

        date_from="",

        date_to=""
    )


# =========================================================
# CASE STATUS FILTER
# =========================================================

@app.route("/status/<status>")
@login_required
def status_cases(status):

    return redirect(
        f"/cases?hearing_status={status}"
    )


# =========================================================
# CASE STAGE FILTER
# =========================================================

@app.route("/stage/<stage>")
@login_required
def stage_cases(stage):

    return redirect(
        f"/cases?case_stage={stage}"
    )


# =========================================================
# ADD CASE
# =========================================================

@app.route(
    "/add-case",
    methods=["GET", "POST"]
)

@login_required
def add_case():

    if request.method == "POST":

        hearing_date = request.form.get(
            "next_hearing_date"
        )
        decision_date_value = request.form.get("decision_date")
        decision_date = datetime.strptime(decision_date_value, "%Y-%m-%d").date() if decision_date_value else None

        new_case = Case(

            case_no=request.form.get(
                "case_no"
            ),

            crn_no=request.form.get(
                "crn_no"
            ) or None,

            advocate_name=request.form.get(
                "advocate_name"
            ) or None,

            case_disposed="Yes" if decision_date else (request.form.get(
                "case_disposed"
            ) or "No"),

            decision_date=decision_date,

            court_no=int(
                request.form.get(
                    "court_no"
                )
            ),

            parties=request.form.get(
                "parties"
            ),

            case_stage=request.form.get(
                "case_stage"
            ),

            case_stage_other=(
                request.form.get(
                    "case_stage_other"
                ).strip()
                if request.form.get(
                    "case_stage"
                ) == "Other"
                and request.form.get(
                    "case_stage_other"
                )
                else None
            ),

            next_hearing_date=(

                datetime.strptime(
                    hearing_date,
                    "%Y-%m-%d"
                ).date()

                if hearing_date

                else None
            )
        )


        db.session.add(
            new_case
        )

        db.session.commit()


        return redirect(
            "/add-case"
        )


    return render_template(
        "add_case.html"
    )


# =========================================================
# EDIT CASE
# =========================================================

@app.route(
    "/edit-case/<int:id>",
    methods=["GET", "POST"]
)

@login_required
def edit_case(id):

    case = Case.query.get_or_404(id)


    if request.method == "POST":

        hearing_date = request.form.get(
            "next_hearing_date"
        )
        decision_date_value = request.form.get("decision_date")
        decision_date = datetime.strptime(decision_date_value, "%Y-%m-%d").date() if decision_date_value else None


        case.case_no = request.form.get(
            "case_no"
        )

        case.crn_no = request.form.get(
            "crn_no"
        ) or None

        case.advocate_name = request.form.get(
            "advocate_name"
        ) or None

        case.decision_date = decision_date
        case.case_disposed = "Yes" if decision_date else (request.form.get(
            "case_disposed"
        ) or "No")

        case.court_no = int(
            request.form.get(
                "court_no"
            )
        )

        case.parties = request.form.get(
            "parties"
        )

        case.case_stage = request.form.get(
            "case_stage"
        )

        case.case_stage_other = (
            request.form.get(
                "case_stage_other"
            ).strip()
            if request.form.get(
                "case_stage"
            ) == "Other"
            and request.form.get(
                "case_stage_other"
            )
            else None
        )

        case.next_hearing_date = (

            datetime.strptime(
                hearing_date,
                "%Y-%m-%d"
            ).date()

            if hearing_date

            else None
        )


        db.session.commit()


        return redirect(
            f"/case/{case.id}"
        )


    return render_template(

        "edit_case.html",

        case=case
    )


# =========================================================
# CASE DETAIL PAGE
# =========================================================

@app.route("/case/<int:id>")
@login_required
def case_detail(id):

    case = Case.query.get_or_404(id)


    hearings = Hearing.query.filter_by(
        case_id=case.id
    ).order_by(
        Hearing.hearing_date.desc()
    ).all()


    evidences = Evidence.query.filter_by(
        case_id=case.id
    ).order_by(
        Evidence.id.desc()
    ).all()


    return render_template(

        "case_detail.html",

        case=case,

        hearings=hearings,

        evidences=evidences,

        today=date.today()
    )


# =========================================================
# VIEW BUSINESS / HEARING DETAIL
# =========================================================

@app.route("/case/<int:case_id>/business/<int:hearing_id>")
@login_required
def view_business(case_id, hearing_id):

    case = Case.query.get_or_404(case_id)
    hearing = Hearing.query.filter_by(
        id=hearing_id,
        case_id=case.id
    ).first_or_404()

    return render_template(
        "view_business.html",
        case=case,
        hearing=hearing
    )


# =========================================================
# EDIT / DELETE CASE HISTORY
# =========================================================

@app.route("/case/<int:case_id>/hearing/<int:hearing_id>/edit", methods=["GET", "POST"])
@login_required
def edit_hearing(case_id, hearing_id):
    case = Case.query.get_or_404(case_id)
    hearing = Hearing.query.filter_by(id=hearing_id, case_id=case.id).first_or_404()

    if request.method == "POST":
        hearing_date_value = request.form.get("hearing_date", "").strip()
        next_date_value = request.form.get("next_hearing_date", "").strip()

        try:
            hearing.hearing_date = datetime.strptime(hearing_date_value, "%Y-%m-%d").date() if hearing_date_value else hearing.hearing_date
            hearing.next_hearing_date = datetime.strptime(next_date_value, "%Y-%m-%d").date() if next_date_value else None
        except ValueError:
            return "Invalid date.", 400

        hearing.outcome = request.form.get("outcome", "").strip() or "Hearing"
        hearing.presentee = request.form.get("presentee", "").strip() or None
        hearing.business = request.form.get("business", "").strip() or None
        hearing.notes = request.form.get("notes", "").strip() or None

        latest = Hearing.query.filter_by(case_id=case.id).order_by(Hearing.hearing_date.desc(), Hearing.id.desc()).first()
        if latest and latest.id == hearing.id:
            case.next_hearing_date = hearing.next_hearing_date

        db.session.commit()
        return redirect(url_for("case_detail", id=case.id))

    return render_template("edit_hearing.html", case=case, hearing=hearing)


@app.route("/case/<int:case_id>/hearing/<int:hearing_id>/delete", methods=["POST"])
@login_required
def delete_hearing(case_id, hearing_id):
    case = Case.query.get_or_404(case_id)
    hearing = Hearing.query.filter_by(id=hearing_id, case_id=case.id).first_or_404()

    db.session.delete(hearing)
    db.session.flush()

    latest = Hearing.query.filter_by(case_id=case.id).order_by(Hearing.hearing_date.desc(), Hearing.id.desc()).first()
    if latest:
        case.next_hearing_date = latest.next_hearing_date
    else:
        case.next_hearing_date = None

    db.session.commit()
    return redirect(url_for("case_detail", id=case.id))


# =========================================================
# QUICK HEARING UPDATE
# =========================================================

@app.route(
    "/case/<int:id>/update-hearing",
    methods=["POST"]
)

@login_required
def update_hearing(id):

    case = Case.query.get_or_404(id)


    outcome = request.form.get("outcome", "").strip()
    if outcome == "Other":
        outcome = request.form.get("other_outcome", "").strip() or "Other"

    hearing_date = request.form.get(
        "hearing_date"
    )


    next_hearing_date = request.form.get(
        "next_hearing_date"
    )


    notes = request.form.get(
        "notes"
    )

    presentee = request.form.get(
        "presentee",
        ""
    ).strip()

    business = request.form.get(
        "business",
        ""
    ).strip()


    if hearing_date:

        hearing_date = datetime.strptime(
            hearing_date,
            "%Y-%m-%d"
        ).date()

    else:

        hearing_date = date.today()


    if next_hearing_date:

        next_hearing_date = datetime.strptime(
            next_hearing_date,
            "%Y-%m-%d"
        ).date()

    else:

        next_hearing_date = None


    new_hearing = Hearing(

        case_id=case.id,

        hearing_date=hearing_date,

        outcome=outcome,

        presentee=presentee or None,

        business=business or None,

        next_hearing_date=next_hearing_date,

        notes=notes
    )


    db.session.add(
        new_hearing
    )


    case.next_hearing_date = (
        next_hearing_date
    )


    if outcome == "Evidence Submitted":

        case.case_stage = "Evidence"

    elif outcome == "Arguments":

        case.case_stage = "Arguments"

    elif outcome == "Order Passed":

        case.case_stage = "Order"

    elif outcome == "Hearing":

        case.case_stage = "Hearing"


    db.session.commit()


    return redirect(
        f"/case/{case.id}"
    )


# =========================================================
# ADD EVIDENCE + MULTIPLE FILE UPLOAD
# =========================================================

@app.route(
    "/case/<int:id>/add-evidence",
    methods=["POST"]
)

@login_required
def add_evidence(id):

    case = Case.query.get_or_404(id)


    submitted_date = request.form.get(
        "submitted_date"
    )


    if submitted_date:

        submitted_date = datetime.strptime(
            submitted_date,
            "%Y-%m-%d"
        ).date()

    else:

        submitted_date = None


    # CREATE EVIDENCE FIRST

    new_evidence = Evidence(

        case_id=case.id,

        evidence_name=request.form.get(
            "evidence_name"
        ),

        evidence_type=request.form.get(
            "evidence_type"
        ),

        description=request.form.get(
            "description"
        ),

        submitted_date=submitted_date,

        status=request.form.get(
            "status"
        )
    )


    db.session.add(
        new_evidence
    )

    db.session.flush()


    # MULTIPLE FILES

    uploaded_files = request.files.getlist(
        "evidence_files"
    )


    for uploaded_file in uploaded_files:


        if not uploaded_file:

            continue


        if uploaded_file.filename == "":

            continue


        if not allowed_file(
            uploaded_file.filename
        ):

            continue


        original_filename = secure_filename(
            uploaded_file.filename
        )


        unique_filename = (

            str(uuid.uuid4())

            + "_"

            + original_filename
        )


        uploaded_file.save(

            os.path.join(

                app.config["UPLOAD_FOLDER"],

                unique_filename
            )
        )


        new_file = EvidenceFile(

            evidence_id=new_evidence.id,

            file_name=unique_filename,

            original_filename=original_filename
        )


        db.session.add(
            new_file
        )


    if request.form.get(
        "status"
    ) == "Submitted":

        case.case_stage = "Evidence"


    db.session.commit()


    return redirect(
        f"/case/{case.id}"
    )


# =========================================================
# VIEW EVIDENCE FILE
# =========================================================

@app.route(
    "/evidence-file/<int:id>"
)

@login_required
def view_evidence_file(id):

    evidence_file = EvidenceFile.query.get_or_404(
        id
    )


    return send_from_directory(

        app.config["UPLOAD_FOLDER"],

        evidence_file.file_name,

        as_attachment=False
    )


# =========================================================
# DELETE SINGLE EVIDENCE FILE
# =========================================================

@app.route(
    "/evidence-file/<int:id>/delete",
    methods=["POST"]
)

@login_required
def delete_evidence_file(id):

    evidence_file = EvidenceFile.query.get_or_404(
        id
    )


    evidence_id = evidence_file.evidence_id


    file_path = os.path.join(

        app.config["UPLOAD_FOLDER"],

        evidence_file.file_name
    )


    if os.path.exists(
        file_path
    ):

        os.remove(
            file_path
        )


    db.session.delete(
        evidence_file
    )

    db.session.commit()


    evidence = Evidence.query.get(
        evidence_id
    )


    return redirect(
        f"/case/{evidence.case_id}"
    )


# =========================================================
# EDIT EVIDENCE
# =========================================================

@app.route(
    "/evidence/<int:id>/edit",
    methods=["GET", "POST"]
)

@login_required
def edit_evidence(id):

    evidence = Evidence.query.get_or_404(id)


    if request.method == "POST":

        submitted_date = request.form.get(
            "submitted_date"
        )


        evidence.evidence_name = request.form.get(
            "evidence_name"
        )

        evidence.evidence_type = request.form.get(
            "evidence_type"
        )

        evidence.description = request.form.get(
            "description"
        )

        evidence.status = request.form.get(
            "status"
        )


        if submitted_date:

            evidence.submitted_date = (
                datetime.strptime(
                    submitted_date,
                    "%Y-%m-%d"
                ).date()
            )

        else:

            evidence.submitted_date = None


        # ADD MORE FILES

        uploaded_files = request.files.getlist(
            "evidence_files"
        )


        for uploaded_file in uploaded_files:


            if not uploaded_file:

                continue


            if uploaded_file.filename == "":

                continue


            if not allowed_file(
                uploaded_file.filename
            ):

                continue


            original_filename = secure_filename(
                uploaded_file.filename
            )


            unique_filename = (

                str(uuid.uuid4())

                + "_"

                + original_filename
            )


            uploaded_file.save(

                os.path.join(

                    app.config["UPLOAD_FOLDER"],

                    unique_filename
                )
            )


            new_file = EvidenceFile(

                evidence_id=evidence.id,

                file_name=unique_filename,

                original_filename=original_filename
            )


            db.session.add(
                new_file
            )


        db.session.commit()


        return redirect(
            f"/case/{evidence.case_id}"
        )


    return render_template(

        "edit_evidence.html",

        evidence=evidence
    )


# =========================================================
# DELETE EVIDENCE + ALL FILES
# =========================================================

@app.route(
    "/evidence/<int:id>/delete",
    methods=["POST"]
)

@login_required
def delete_evidence(id):

    evidence = Evidence.query.get_or_404(id)


    case_id = evidence.case_id


    for evidence_file in evidence.files:


        file_path = os.path.join(

            app.config["UPLOAD_FOLDER"],

            evidence_file.file_name
        )


        if os.path.exists(
            file_path
        ):

            os.remove(
                file_path
            )


    db.session.delete(
        evidence
    )

    db.session.commit()


    return redirect(
        f"/case/{case_id}"
    )


# =========================================================
# IMPORT CASES FROM EXCEL
# =========================================================

@app.route(
    "/import-excel",
    methods=["GET", "POST"]
)

@login_required
def import_excel():

    if request.method == "POST":

        excel_file = request.files.get(
            "excel_file"
        )

        if not excel_file:
            return "Please select an Excel file."

        if not excel_file.filename.lower().endswith(
            ".xlsx"
        ):
            return "Please upload a valid .xlsx file."

        try:

            workbook = load_workbook(
                excel_file,
                data_only=True
            )

            worksheet = workbook.active

            # Header-based import allows flexible column order.
            headers = {}

            for index, cell in enumerate(
                worksheet[1]
            ):

                if cell.value is not None:

                    headers[
                        str(cell.value).strip().lower()
                    ] = index


            def get_value(row, *names):

                for name in names:

                    key = name.lower()

                    if key in headers:

                        return row[
                            headers[key]
                        ]

                return None


            for row in worksheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                case_no = get_value(
                    row,
                    "Case Number",
                    "Case No",
                    "case_no"
                )

                crn_no = get_value(
                    row,
                    "CRN",
                    "CRN No",
                    "crn_no"
                )

                court_no = get_value(
                    row,
                    "Court Number",
                    "Court No",
                    "court_no"
                )

                parties = get_value(
                    row,
                    "Name of Parties",
                    "Parties",
                    "parties"
                )

                advocate_name = get_value(
                    row,
                    "Name of Advocate",
                    "Advocate Name",
                    "advocate_name"
                )

                case_stage = get_value(
                    row,
                    "Case Stage",
                    "case_stage"
                )

                case_stage_other = get_value(
                    row,
                    "Case Stage Other",
                    "Other Case Stage",
                    "case_stage_other"
                )

                case_disposed = get_value(
                    row,
                    "Case Disposed",
                    "Disposed",
                    "case_disposed"
                )

                hearing_date = get_value(
                    row,
                    "Next Hearing Date",
                    "Hearing Date",
                    "next_hearing_date"
                )

                if not all([
                    case_no,
                    court_no,
                    parties,
                    case_stage
                ]):
                    continue

                try:
                    court_no = int(court_no)
                except (TypeError, ValueError):
                    continue

                if isinstance(
                    hearing_date,
                    datetime
                ):
                    hearing_date = hearing_date.date()

                elif isinstance(
                    hearing_date,
                    str
                ) and hearing_date.strip():

                    parsed_date = None

                    for date_format in [
                        "%Y-%m-%d",
                        "%d-%m-%Y",
                        "%d/%m/%Y"
                    ]:

                        try:
                            parsed_date = datetime.strptime(
                                hearing_date.strip(),
                                date_format
                            ).date()

                            break

                        except ValueError:
                            pass

                    hearing_date = parsed_date

                else:
                    hearing_date = None


                new_case = Case(

                    case_no=str(
                        case_no
                    ).strip(),

                    crn_no=(
                        str(crn_no).strip()
                        if crn_no
                        else None
                    ),

                    court_no=court_no,

                    parties=str(
                        parties
                    ).strip(),

                    advocate_name=(
                        str(advocate_name).strip()
                        if advocate_name
                        else None
                    ),

                    case_stage=str(
                        case_stage
                    ).strip(),

                    case_stage_other=(
                        str(case_stage_other).strip()
                        if case_stage_other
                        else None
                    ),

                    case_disposed=(
                        str(case_disposed).strip()
                        if case_disposed
                        else "No"
                    ),

                    next_hearing_date=hearing_date
                )

                db.session.add(
                    new_case
                )


            db.session.commit()

            return redirect(
                "/cases"
            )

        except Exception as error:

            db.session.rollback()

            return f"Error importing file: {error}"


    return render_template(
        "import_excel.html"
    )


# =========================================================
# EXPORT CASES TO EXCEL
# =========================================================

@app.route("/export-excel")
@login_required
def export_excel():

    cases = sort_cases_naturally(Case.query.all())


    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Court Cases"


    worksheet.append([

        "ID",
        "Case Number",
        "CRN",
        "Court Number",
        "Name of Parties",
        "Name of Advocate",
        "Case Stage",
        "Case Stage Other",
        "Case Disposed",
        "Next Hearing Date"
    ])


    for case in cases:


        worksheet.append([

            case.id,

            case.case_no,

            case.crn_no or "",

            case.court_no,

            case.parties,

            case.advocate_name or "",

            case.case_stage,

            case.case_stage_other or "",

            case.case_disposed or "No",

            case.next_hearing_date.strftime(
                "%d-%m-%Y"
            )

            if case.next_hearing_date

            else ""
        ])


    excel_file = BytesIO()

    workbook.save(
        excel_file
    )

    excel_file.seek(
        0
    )


    return send_file(

        excel_file,

        as_attachment=True,

        download_name="court_cases.xlsx",

        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


# =========================================================
# DELETE CASE + ALL EVIDENCE FILES
# =========================================================

@app.route(
    "/delete-case/<int:id>",
    methods=["POST"]
)

@login_required
def delete_case(id):

    case = Case.query.get_or_404(id)


    for evidence in case.evidences:

        for evidence_file in evidence.files:


            file_path = os.path.join(

                app.config["UPLOAD_FOLDER"],

                evidence_file.file_name
            )


            if os.path.exists(
                file_path
            ):

                os.remove(
                    file_path
                )


    db.session.delete(
        case
    )

    db.session.commit()


    return redirect(
        "/cases"
    )


# =========================================================
# SQLITE DATABASE MIGRATION FOR NEW CASE FIELDS
# =========================================================

def create_default_admin():

    admin = User.query.filter_by(
        username="admin"
    ).first()

    if not admin:

        admin = User(
            username="admin",
            password_hash=generate_password_hash("admin")
        )

        db.session.add(admin)

    else:

        admin.password_hash = generate_password_hash(
            "admin"
        )

    db.session.commit()

def add_missing_case_columns():

    inspector = db.inspect(db.engine)
    existing_names = {column["name"] for column in inspector.get_columns("case")}

    columns_to_add = {
        "crn_no": "VARCHAR(100)",
        "advocate_name": "VARCHAR(200)",
        "case_disposed": "VARCHAR(20) DEFAULT 'No'",
        "case_stage_other": "VARCHAR(200)",
        "decision_date": "DATE",
        "highlighted": "BOOLEAN DEFAULT FALSE"
    }

    for column_name, column_type in columns_to_add.items():

        if column_name not in existing_names:

            db.session.execute(
                db.text(
                    f'ALTER TABLE "case" ADD COLUMN '
                    f"{column_name} {column_type}"
                )
            )

    db.session.commit()

# =========================================================
# HEARING TABLE MIGRATION
# =========================================================

def add_missing_hearing_columns():

    inspector = db.inspect(db.engine)
    table_names = inspector.get_table_names()

    if "hearing" not in table_names:
        return

    existing_names = {column["name"] for column in inspector.get_columns("hearing")}

    columns_to_add = {
        "presentee": "TEXT",
        "business": "TEXT"
    }

    for column_name, column_type in columns_to_add.items():
        if column_name not in existing_names:
            db.session.execute(
                db.text(
                    f'ALTER TABLE "hearing" ADD COLUMN {column_name} {column_type}'
                )
            )

    db.session.commit()

# =========================================================
# CREATE DATABASE TABLES AND RUN APP
# =========================================================


# =========================================================
# BULK ACTIONS + ANALYTICS
# =========================================================
def _bulk_cases():
    raw = request.args.get("ids", "")
    ids = []
    for x in raw.split(","):
        try: ids.append(int(x))
        except ValueError: pass
    return Case.query.filter(Case.id.in_(ids)).all()

@app.route("/bulk/date", methods=["POST"])
@login_required
def bulk_date():
    value=request.form.get("next_hearing_date","")
    if not value: flash("Please select a date.","warning"); return redirect(request.referrer or url_for("case_list"))
    try: d=datetime.strptime(value,"%Y-%m-%d").date()
    except ValueError: return "Invalid date",400
    for c in _bulk_cases(): c.next_hearing_date=d
    db.session.commit(); flash("Next hearing date updated for selected cases.","success")
    return redirect(request.referrer or url_for("case_list"))

@app.route("/bulk/stage", methods=["POST"])
@login_required
def bulk_stage():
    stage=request.form.get("case_stage","").strip()
    if not stage: flash("Please select a stage.","warning"); return redirect(request.referrer or url_for("case_list"))
    for c in _bulk_cases(): c.case_stage=stage
    db.session.commit(); flash("Case stage updated for selected cases.","success")
    return redirect(request.referrer or url_for("case_list"))

@app.route("/bulk/dispose", methods=["POST"])
@login_required
def bulk_dispose():
    value=request.form.get("decision_date","")
    if not value: return "Decision date is required.",400
    try: d=datetime.strptime(value,"%Y-%m-%d").date()
    except ValueError: return "Invalid date",400
    for c in _bulk_cases():
        c.case_disposed="Yes"; c.decision_date=d; c.next_hearing_date=None
    db.session.commit(); flash("Selected cases disposed.","success")
    return redirect(request.referrer or url_for("case_list"))

@app.route("/bulk/delete", methods=["POST"])
@login_required
def bulk_delete():
    for c in _bulk_cases(): db.session.delete(c)
    db.session.commit(); flash("Selected cases deleted.","success")
    return redirect(request.referrer or url_for("case_list"))

# =========================================================
# ENQUIRIES - INDEPENDENT FROM CASES
# =========================================================

ENQUIRY_STATUSES = ["Pending for Report", "Report Drafting", "Report Given"]

def _enquiry_status_options():
    """Return fixed statuses plus every custom status already saved in the database."""
    fixed = list(ENQUIRY_STATUSES)
    custom = set()
    for row in db.session.query(Enquiry.status).filter(Enquiry.status.isnot(None)).all():
        value = (row[0] or "").strip()
        if value and value not in fixed and value not in ("Other", "__other__"):
            custom.add(value)
    return fixed + sorted(custom, key=str.casefold)

def _ensure_enquiry_tables():
    """Create Enquiry tables if this deployment database does not have them yet."""
    db.create_all()

def _parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None

def _parse_time(value):
    return datetime.strptime(value, "%H:%M").time() if value else None

@app.route("/enquiries")
@login_required
def enquiry_list():
    _ensure_enquiry_tables()
    search = request.args.get("search", "").strip()
    status = request.args.get("status", "").strip()
    highlighted_only = request.args.get("highlighted", "") == "1"
    schedule = request.args.get("schedule", "").strip()
    today = date.today()
    query = Enquiry.query
    if search:
        query = query.filter(Enquiry.organisation_name.ilike(f"%{search}%"))
    if status:
        query = query.filter(Enquiry.status == status)
    if highlighted_only:
        query = query.filter(Enquiry.highlighted.is_(True))
    if schedule == "today":
        query = query.filter(Enquiry.next_enquiry_date == today)
    elif schedule == "next":
        query = query.filter(Enquiry.next_enquiry_date > today)
    enquiries = query.order_by(Enquiry.next_enquiry_date.is_(None), Enquiry.next_enquiry_date, Enquiry.id.desc()).all()
    status_options = _enquiry_status_options()
    return render_template("enquiries.html", enquiries=enquiries, statuses=status_options,
                           selected_status=status, search=search, highlighted_only=highlighted_only, today=date.today())

@app.route("/enquiry/add", methods=["GET", "POST"])
@login_required
def add_enquiry():
    _ensure_enquiry_tables()
    if request.method == "POST":
        name = request.form.get("organisation_name", "").strip()
        start = _parse_date(request.form.get("start_date", ""))
        next_date = _parse_date(request.form.get("next_enquiry_date", ""))
        next_time = _parse_time(request.form.get("next_enquiry_time", ""))
        status = request.form.get("status", "Pending for Report").strip()
        custom_status = request.form.get("custom_status", "").strip()
        if status in ("Other", "__other__") and custom_status:
            status = custom_status
        if not name or not start:
            flash("Organisation name and start date are required.", "warning")
            return redirect(url_for("add_enquiry"))
        if not status or (status in ("Other", "__other__") and not custom_status):
            status = "Pending for Report"
        enquiry = Enquiry(organisation_name=name, start_date=start, next_enquiry_date=next_date,
                          next_enquiry_time=next_time, status=status)
        db.session.add(enquiry)
        db.session.flush()
        # The opening date is counted as the first sitting/turn.
        db.session.add(EnquirySitting(enquiry_id=enquiry.id, sitting_date=start, sitting_time=next_time, remark="Enquiry started"))
        db.session.commit()
        flash("Enquiry added successfully.", "success")
        return redirect(url_for("enquiry_list"))
    return render_template("add_enquiry.html", statuses=_enquiry_status_options(), today=date.today())

@app.route("/enquiry/<int:id>")
@login_required
def enquiry_detail(id):
    _ensure_enquiry_tables()
    enquiry = Enquiry.query.get_or_404(id)
    sittings = EnquirySitting.query.filter_by(enquiry_id=id).order_by(EnquirySitting.sitting_date.desc(), EnquirySitting.id.desc()).all()
    return render_template("enquiry_detail.html", enquiry=enquiry, sittings=sittings, statuses=_enquiry_status_options(), today=date.today())

@app.route("/enquiry/<int:id>/edit", methods=["GET", "POST"])
@login_required
def edit_enquiry(id):
    _ensure_enquiry_tables()
    enquiry = Enquiry.query.get_or_404(id)
    if request.method == "POST":
        name = request.form.get("organisation_name", "").strip()
        start = _parse_date(request.form.get("start_date", ""))
        end = _parse_date(request.form.get("end_date", ""))
        next_date = _parse_date(request.form.get("next_enquiry_date", ""))
        next_time = _parse_time(request.form.get("next_enquiry_time", ""))
        status = request.form.get("status", enquiry.status).strip()
        custom_status = request.form.get("custom_status", "").strip()
        if status in ("Other", "__other__") and custom_status:
            status = custom_status
        if not name or not start:
            flash("Organisation name and start date are required.", "warning")
            return redirect(url_for("edit_enquiry", id=id))
        enquiry.organisation_name=name; enquiry.start_date=start; enquiry.end_date=end
        enquiry.next_enquiry_date=next_date; enquiry.next_enquiry_time=next_time
        enquiry.status = status if status else enquiry.status
        db.session.commit()
        flash("Enquiry updated successfully.", "success")
        return redirect(url_for("enquiry_detail", id=id))
    return render_template("add_enquiry.html", enquiry=enquiry, statuses=_enquiry_status_options(), today=date.today(), edit_mode=True)

@app.route("/enquiry/<int:id>/add-sitting", methods=["POST"])
@login_required
def add_enquiry_sitting(id):
    _ensure_enquiry_tables()
    enquiry = Enquiry.query.get_or_404(id)
    sitting_date = _parse_date(request.form.get("sitting_date", ""))
    sitting_time = _parse_time(request.form.get("sitting_time", ""))
    remark = request.form.get("remark", "").strip()
    next_date = _parse_date(request.form.get("next_enquiry_date", ""))
    next_time = _parse_time(request.form.get("next_enquiry_time", ""))
    if not sitting_date:
        flash("Sitting date is required.", "warning")
        return redirect(url_for("enquiry_detail", id=id))
    db.session.add(EnquirySitting(enquiry_id=id, sitting_date=sitting_date, sitting_time=sitting_time, remark=remark))
    enquiry.next_enquiry_date=next_date; enquiry.next_enquiry_time=next_time
    db.session.commit()
    flash("Enquiry sitting added. Total turns updated automatically.", "success")
    return redirect(url_for("enquiry_detail", id=id))

@app.route("/enquiry/<int:id>/toggle-highlight", methods=["POST"])
@login_required
def toggle_enquiry_highlight(id):
    _ensure_enquiry_tables()
    enquiry=Enquiry.query.get_or_404(id)
    enquiry.highlighted=not bool(enquiry.highlighted)
    db.session.commit()
    return redirect(request.referrer or url_for("enquiry_list"))

@app.route("/enquiry/<int:id>/delete", methods=["POST"])
@login_required
def delete_enquiry(id):
    _ensure_enquiry_tables()
    enquiry=Enquiry.query.get_or_404(id)
    db.session.delete(enquiry); db.session.commit()
    flash("Enquiry deleted.", "success")
    return redirect(url_for("enquiry_list"))

@app.route("/analytics")
@login_required
def analytics():
    today=date.today()
    total=Case.query.count()
    active=Case.query.filter(Case.case_disposed != "Yes").count()
    disposed=Case.query.filter_by(case_disposed="Yes").count()
    undated=Case.query.filter(Case.next_hearing_date.is_(None),Case.case_disposed!="Yes").count()
    next7=Case.query.filter(Case.next_hearing_date>=today,Case.next_hearing_date<=today+timedelta(days=7)).count()
    days=["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday"]
    counts=[]
    for i in range(6):
        counts.append(Case.query.filter(db.extract("dow", Case.next_hearing_date)==i).count())
    # PostgreSQL dow differs; calculate reliably in Python from scheduled cases.
    counts=[0]*6
    for c in Case.query.filter(Case.next_hearing_date.isnot(None)).all():
        if c.next_hearing_date.weekday()<6: counts[c.next_hearing_date.weekday()]+=1
    courts=[Case.query.filter_by(court_no=n).count() for n in (1,2,3)]
    mx=max(counts) if counts else 0
    positive=[x for x in counts if x>0]
    mn=min(positive) if positive else 0
    return render_template("analytics.html",total=total,active=active,disposed=disposed,undated=undated,next7=next7,days=days,counts=counts,courts=courts,busiest=days[counts.index(mx)] if counts else "-",busiest_count=mx,least=days[counts.index(mn)] if positive else "-",least_count=mn,no_days=[days[i] for i,x in enumerate(counts) if x==0])


with app.app_context():
    db.create_all()
    add_missing_case_columns()
    add_missing_hearing_columns()
    create_default_admin()


if __name__ == "__main__":
    app.run(
        debug=os.environ.get("FLASK_DEBUG") == "1",
        host="0.0.0.0",
        port=int(os.environ.get("PORT", 5000))
    )
